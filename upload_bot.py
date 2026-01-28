from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from dotenv import load_dotenv
from datetime import datetime
from docx import Document
from utils import extract_content_sequence, ensure_english_filenames, sanitize_filename
import zipfile
import pyperclip
import pyautogui
import platform
import time
import os
import glob
import re
import tempfile

pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.3

# .env 파일에서 환경 변수 로드
load_dotenv()

# 네이버 계정 정보 (.env 파일에서 읽어오기)
NAVER_ID = os.getenv("NAVER_ID")
NAVER_PW = os.getenv("NAVER_PW")

# 계정 정보 확인
if not NAVER_ID or not NAVER_PW:
    raise ValueError(".env 파일에서 NAVER_ID 또는 NAVER_PW를 찾을 수 없습니다.")

# OS에 따른 붙여넣기 키 설정 (Mac: Command, Windows/Linux: Control)
PASTE_KEY = Keys.COMMAND if platform.system() == 'Darwin' else Keys.CONTROL

# Chrome 옵션 설정
chrome_options = Options()
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-dev-shm-usage')

# WebDriver 초기화
driver = webdriver.Chrome(options=chrome_options)

def find_blog_excel_file():
    """
    현재 디렉토리에서 blog+날짜.xlsx 파일을 찾습니다.
    
    Returns:
        str: 파일 경로
    """
    current_dir = os.getcwd()
    # blog로 시작하고 .xlsx로 끝나는 파일 찾기
    pattern = os.path.join(current_dir, "blog*.xlsx")
    files = glob.glob(pattern)
    
    if not files:
        raise FileNotFoundError("blog+날짜.xlsx 파일을 찾을 수 없습니다.")
    
    # 가장 최근 파일 반환 (여러 개 있을 경우)
    latest_file = max(files, key=os.path.getmtime)
    return latest_file

def naver_login():
    print("네이버 로그인 페이지 접속 중...")
    driver.get("https://nid.naver.com/nidlogin.login")
    time.sleep(3)
    
    print("아이디 입력 중...")
    id_input = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "id"))
    )
    id_input.click()
    time.sleep(0.5)
    pyperclip.copy(NAVER_ID)
    id_input.send_keys(PASTE_KEY, 'v')
    time.sleep(1)
    
    current_id = id_input.get_attribute('value')
    if not current_id:
        print("  클립보드 방식 실패, JavaScript 방식 시도...")
        driver.execute_script(f"arguments[0].value = '{NAVER_ID}';", id_input)
        driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", id_input)
        time.sleep(0.5)
    
    print("비밀번호 입력 중...")
    pw_input = driver.find_element(By.ID, "pw")
    pw_input.click()
    time.sleep(0.5)
    pyperclip.copy(NAVER_PW)
    pw_input.send_keys(PASTE_KEY, 'v')
    time.sleep(1)
    
    current_pw = pw_input.get_attribute('value')
    if not current_pw:
        print("  클립보드 방식 실패, JavaScript 방식 시도...")
        driver.execute_script(f"arguments[0].value = '{NAVER_PW}';", pw_input)
        driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", pw_input)
        time.sleep(0.5)
    
    print("로그인 버튼 클릭 중...")
    login_button = driver.find_element(By.ID, "log.login")
    login_button.click()
    
    print("로그인 처리 중... (3초 대기)")
    time.sleep(3)
    
    current_url = driver.current_url
    if "nidlogin" not in current_url:
        print("로그인 성공!\n")
    else:
        print("로그인 실패 - 수동 확인 필요\n")

def select_category(category_name):
    """발행 레이어에서 카테고리 선택"""
    if not category_name:
        return
    
    print(f"  카테고리 선택 중: {category_name}")
    try:
        cat_button = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".selectbox_button__jb1Dt"))
        )
        driver.execute_script("arguments[0].click();", cat_button)
        time.sleep(1)
        
        cat_labels = driver.find_elements(By.CSS_SELECTOR, ".option_list_layer__YX1Tq label.radio_label__mB6ia")
        for label in cat_labels:
            label_text = label.text.strip()
            if category_name in label_text or label_text in category_name:
                driver.execute_script("arguments[0].click();", label)
                print(f"    - 카테고리 '{label_text}' 선택 완료")
                time.sleep(0.5)
                return
        
        print(f"    - 카테고리 '{category_name}'를 찾을 수 없음")
        driver.execute_script("arguments[0].click();", cat_button)
    except Exception as e:
        print(f"    - 카테고리 선택 실패: {e}")


def set_schedule_time(schedule_time_str):
    """예약 발행 시간 설정 (형식: YYYY-MM-DD HH:MM)"""
    if not schedule_time_str:
        return
    
    print(f"  예약 발행 설정 중: {schedule_time_str}")
    try:
        from selenium.webdriver.support.ui import Select
        
        parts = schedule_time_str.strip().split()
        if len(parts) != 2:
            print(f"    - 잘못된 형식 (YYYY-MM-DD HH:MM 필요)")
            return
        
        date_part, time_part = parts
        hour, minute = time_part.split(":")
        
        minute_rounded = str((int(minute) // 10) * 10).zfill(2)
        
        reserve_radio = driver.find_element(By.CSS_SELECTOR, "input#radio_time2[value='pre']")
        driver.execute_script("arguments[0].click();", reserve_radio)
        time.sleep(1)
        
        date_input = driver.find_element(By.CSS_SELECTOR, "input.input_date__QmA0s")
        driver.execute_script("arguments[0].click();", date_input)
        time.sleep(1)
        
        try:
            year, month, day = date_part.split("-")
            day_buttons = driver.find_elements(By.CSS_SELECTOR, ".react-datepicker__day:not(.react-datepicker__day--outside-month)")
            for btn in day_buttons:
                if btn.text.strip() == str(int(day)):
                    driver.execute_script("arguments[0].click();", btn)
                    print(f"    - 날짜 {day}일 선택")
                    break
            time.sleep(0.5)
        except:
            print(f"    - 날짜 선택 실패, 기본 날짜 사용")
        
        hour_select = Select(driver.find_element(By.CSS_SELECTOR, "select.hour_option__J_heO"))
        hour_select.select_by_value(hour.zfill(2))
        print(f"    - 시간 {hour}시 선택")
        
        minute_select = Select(driver.find_element(By.CSS_SELECTOR, "select.minute_option__Vb3xB"))
        minute_select.select_by_value(minute_rounded)
        print(f"    - 분 {minute_rounded}분 선택")
        
        time.sleep(0.5)
        print(f"    - 예약 발행 설정 완료")
    except Exception as e:
        print(f"    - 예약 발행 설정 실패: {e}")


def copy_image_to_clipboard(image_path):
    import subprocess
    
    if platform.system() == 'Darwin':
        script = f'''
        use framework "AppKit"
        use scripting additions
        
        set pb to current application's NSPasteboard's generalPasteboard()
        pb's clearContents()
        
        set img to current application's NSImage's alloc()'s initWithContentsOfFile:"{image_path}"
        pb's writeObjects:{{img}}
        
        return "OK"
        '''
        try:
            result = subprocess.run(['osascript', '-e', script], capture_output=True, text=True)
            return result.returncode == 0
        except:
            return False
    elif platform.system() == 'Windows':
        script = f'''
        Add-Type -AssemblyName System.Windows.Forms
        $image = [System.Drawing.Image]::FromFile("{image_path}")
        [System.Windows.Forms.Clipboard]::SetImage($image)
        '''
        try:
            subprocess.run(['powershell', '-Command', script], check=True, capture_output=True)
            return True
        except:
            return False
    
    return False


def upload_image(image_path):
    if not image_path or not os.path.exists(image_path):
        print(f"  - Image not found: {image_path}")
        return False
    
    abs_path = os.path.abspath(image_path)
    print(f"  Uploading: {os.path.basename(image_path)}")
    
    try:
        content_area = driver.find_element(By.CSS_SELECTOR, ".se-component-content")
        driver.execute_script("arguments[0].click();", content_area)
        time.sleep(0.3)
        
        if copy_image_to_clipboard(abs_path):
            ActionChains(driver).key_down(PASTE_KEY).send_keys('v').key_up(PASTE_KEY).perform()
            time.sleep(3)
            print(f"    - Upload complete (clipboard)")
            return True
        else:
            print(f"    - Clipboard copy failed, trying file dialog...")
            image_btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-name='image']"))
            )
            driver.execute_script("arguments[0].click();", image_btn)
            time.sleep(2)
            
            if platform.system() == 'Darwin':
                pyautogui.hotkey('command', 'shift', 'g')
                time.sleep(1)
                pyperclip.copy(abs_path)
                pyautogui.hotkey('command', 'v')
                time.sleep(0.5)
                pyautogui.press('enter')
                time.sleep(1)
                pyautogui.press('enter')
                time.sleep(3)
            else:
                pyperclip.copy(abs_path)
                pyautogui.hotkey('ctrl', 'v')
                time.sleep(0.5)
                pyautogui.press('enter')
                time.sleep(3)
            
            print(f"    - Upload complete (dialog)")
            return True
    except Exception as e:
        print(f"    - Upload failed: {e}")
        try:
            pyautogui.press('escape')
        except:
            pass
        time.sleep(0.5)
        return False


def write_blog_post(title, content, category=None, schedule_time=None, image_paths=None, content_sequence=None):
    print("Navigate to blog write page...")
    driver.get("https://blog.naver.com/GoBlogWrite.naver")
    time.sleep(2)
    
    print("Switch to iframe...")
    main_frame = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "mainFrame"))
    )
    driver.switch_to.frame(main_frame)
    time.sleep(1)
    
    print("Close popups...")
    for selector in [".se-popup-button-cancel", ".se-help-panel-close-button"]:
        try:
            el = driver.find_element(By.CSS_SELECTOR, selector)
            el.click()
            time.sleep(0.5)
        except:
            pass
    
    print(f"Input title: {title}")
    title_element = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, ".se-section-documentTitle"))
    )
    title_element.click()
    time.sleep(0.5)
    
    title_paragraph = driver.find_element(By.CSS_SELECTOR, ".se-section-documentTitle .se-text-paragraph")
    title_paragraph.click()
    time.sleep(0.3)
    
    pyperclip.copy(str(title))
    ActionChains(driver).key_down(PASTE_KEY).send_keys('v').key_up(PASTE_KEY).perform()
    time.sleep(0.5)
    
    if content_sequence:
        print(f"Input content sequence ({len(content_sequence)} items)...")
        for i, item in enumerate(content_sequence):
            if item["type"] == "text":
                print(f"  [{i+1}] Text: {len(item['content'])} chars")
                text_paragraph = driver.find_element(By.CSS_SELECTOR, ".se-section-text .se-text-paragraph")
                driver.execute_script("arguments[0].click();", text_paragraph)
                time.sleep(0.3)
                pyperclip.copy(item["content"])
                ActionChains(driver).key_down(PASTE_KEY).send_keys('v').key_up(PASTE_KEY).perform()
                ActionChains(driver).send_keys(Keys.ENTER).perform()
                time.sleep(0.5)
            elif item["type"] == "image":
                print(f"  [{i+1}] Image: {os.path.basename(item['path'])}")
                upload_image(item["path"])
                time.sleep(1)
    else:
        print("Input content...")
        content_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".se-section-text"))
        )
        content_element.click()
        time.sleep(0.5)
        
        text_paragraph = driver.find_element(By.CSS_SELECTOR, ".se-section-text .se-text-paragraph")
        text_paragraph.click()
        time.sleep(0.3)
        
        content_str = str(content) if content else ""
        pyperclip.copy(content_str)
        ActionChains(driver).key_down(PASTE_KEY).send_keys('v').key_up(PASTE_KEY).perform()
        time.sleep(1)
        
        if image_paths:
            print(f"Upload images ({len(image_paths)})...")
            for i, img_path in enumerate(image_paths):
                if img_path and os.path.exists(img_path):
                    print(f"  [{i+1}/{len(image_paths)}] {os.path.basename(img_path)}")
                    upload_image(img_path)
                    time.sleep(1)
    
    print("Click publish button...")
    try:
        help_close = driver.find_element(By.CSS_SELECTOR, ".se-help-panel-close-button")
        help_close.click()
        time.sleep(0.5)
    except:
        pass
    
    publish_button = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".publish_btn__m9KHH"))
    )
    driver.execute_script("arguments[0].click();", publish_button)
    time.sleep(2)
    
    print("Configure publish settings...")
    time.sleep(1)
    
    select_category(category)
    set_schedule_time(schedule_time)
    
    print("Confirm publish...")
    try:
        confirm_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button.confirm_btn__WEaBq"))
        )
        driver.execute_script("arguments[0].click();", confirm_button)
    except Exception as e:
        print(f"  - Confirm failed: {e}")
    
    time.sleep(3)
    print("Post published!\n")

def extract_from_word(docx_path, temp_dir):
    doc = Document(docx_path)
    content = '\n'.join([p.text for p in doc.paragraphs if p.text.strip()])
    
    image_paths = []
    try:
        with zipfile.ZipFile(docx_path, 'r') as z:
            for i, name in enumerate(z.namelist()):
                if name.startswith('word/media/'):
                    ext = os.path.splitext(name)[1]
                    img_filename = f"img_{i:03d}{ext}"
                    img_path = os.path.join(temp_dir, img_filename)
                    with z.open(name) as src, open(img_path, 'wb') as dst:
                        dst.write(src.read())
                    image_paths.append(img_path)
    except:
        pass
    
    return content, image_paths


def extract_from_word_sequence(docx_path, temp_dir):
    return extract_content_sequence(docx_path, temp_dir)


def find_word_file(output_dir, title, row_num=None):
    if not os.path.exists(output_dir):
        return None
    
    if row_num is not None:
        new_format = os.path.join(output_dir, f"post_{row_num:03d}.docx")
        if os.path.exists(new_format):
            return new_format
    
    for f in os.listdir(output_dir):
        if f.endswith('.docx') and not f.startswith('test_'):
            name = re.sub(r'^\d+_', '', f.replace('.docx', ''))
            if title in name or name in title:
                return os.path.join(output_dir, f)
    return None


def main():
    try:
        excel_file = find_blog_excel_file()
        print(f"Open excel: {excel_file}")
        wb = load_workbook(excel_file)
        ws = wb.active
        
        if ws is None:
            print("No active worksheet found")
            wb.close()
            return
        
        base_dir = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(base_dir, "output")
        images_dir = os.path.join(base_dir, "images")
        temp_dir = tempfile.mkdtemp()
        
        ensure_english_filenames(output_dir)
        ensure_english_filenames(images_dir)
        
        print(f"Output folder: {output_dir}")
        if os.path.exists(output_dir):
            docx_files = [f for f in os.listdir(output_dir) if f.endswith('.docx')]
            print(f"Word files: {len(docx_files)}")
        
        max_row = ws.max_row or 1
        blog_posts = []
        
        for row in range(2, max_row + 1):
            title = ws[f'A{row}'].value
            category = ws[f'C{row}'].value
            schedule_time = ws[f'D{row}'].value
            
            if not title:
                continue
            
            word_file = find_word_file(output_dir, str(title), row - 1)
            content_sequence = None
            content = None
            image_paths = []
            
            if word_file and os.path.exists(word_file):
                print(f"[Row {row}] Word file: {os.path.basename(word_file)}")
                content_sequence = extract_from_word_sequence(word_file, temp_dir)
                text_items = [item["content"] for item in content_sequence if item["type"] == "text"]
                content = '\n'.join(text_items)
            else:
                content = ws[f'B{row}'].value
                image_paths_str = ws[f'E{row}'].value
                if image_paths_str:
                    for p in str(image_paths_str).split(','):
                        p = p.strip()
                        if os.path.exists(p):
                            image_paths.append(p)
                else:
                    pattern = os.path.join(images_dir, f"section_{row-1}_*.png")
                    found = glob.glob(pattern)
                    if found:
                        image_paths = sorted(found)
            
            if title and (content or content_sequence):
                blog_posts.append((row, title, content, category, schedule_time, image_paths, content_sequence))
        
        if not blog_posts:
            print("No blog posts to process.")
            wb.close()
            return
        
        print(f"Total {len(blog_posts)} posts to upload.\n")
        
        naver_login()
        
        for row, title, content, category, schedule_time, image_paths, content_sequence in blog_posts:
            try:
                print(f"[Row {row}] Start upload")
                print(f"  Title: {title}")
                if category:
                    print(f"  Category: {category}")
                if schedule_time:
                    print(f"  Schedule: {schedule_time}")
                if content_sequence:
                    img_count = len([i for i in content_sequence if i["type"] == "image"])
                    print(f"  Sequence: {len(content_sequence)} items, {img_count} images")
                elif image_paths:
                    print(f"  Images: {len(image_paths)}")
                
                write_blog_post(title, content, category, schedule_time, image_paths, content_sequence)
                
                print(f"[Row {row}] Upload complete\n")
                print("-" * 50 + "\n")
                
            except Exception as e:
                print(f"[Row {row}] Error: {e}")
                import traceback
                traceback.print_exc()
                print("-" * 50 + "\n")
                try:
                    driver.switch_to.default_content()
                except:
                    pass
                continue
        
        wb.close()
        print("All posts uploaded!")
        
    except Exception as e:
        print(f"오류 발생: {e}")
        import traceback
        traceback.print_exc()
    finally:
        # 브라우저 닫기 (필요시 주석 처리)
        # driver.quit()
        pass

if __name__ == "__main__":
    main()
