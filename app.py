import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import threading
import subprocess
import sys
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook
import glob

try:
    from tkcalendar import DateEntry
    HAS_CALENDAR = True
except ImportError:
    HAS_CALENDAR = False

def get_base_path():
    if getattr(sys, 'frozen', False):
        if sys.platform == 'darwin':
            return os.path.dirname(os.path.dirname(sys.executable))
        else:
            return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

def get_python_executable():
    base_path = get_base_path()
    venv_python = os.path.join(base_path, "venv", "bin", "python")
    if os.path.exists(venv_python):
        return venv_python
    venv_python_win = os.path.join(base_path, "venv", "Scripts", "python.exe")
    if os.path.exists(venv_python_win):
        return venv_python_win
    return sys.executable

class BlogAutomationGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("네이버 블로그 자동화 도구")
        self.root.geometry("900x850")
        self.root.resizable(True, True)
        self.root.configure(bg='#f5f7fa')
        
        self.log_file_path = os.path.join(get_base_path(), "app_log.txt")
        self.buttons = []
        
        header_frame = tk.Frame(root, bg='#667eea', height=100)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(
            header_frame, 
            text="네이버 블로그 자동화 도구", 
            font=("Helvetica", 22, "bold"),
            bg='#667eea',
            fg='white'
        )
        title_label.pack(pady=15)
        
        subtitle_label = tk.Label(
            header_frame,
            text="AI 기반 블로그 콘텐츠 자동 생성 및 업로드",
            font=("Helvetica", 11),
            bg='#667eea',
            fg='#e0e7ff'
        )
        subtitle_label.pack(pady=(0, 10))
        
        container = tk.Frame(root, bg='#f5f7fa')
        container.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        button_card = tk.Frame(container, bg='white', relief=tk.FLAT, bd=0)
        button_card.pack(fill=tk.X, pady=(0, 12))
        
        button_inner = tk.Frame(button_card, bg='white')
        button_inner.pack(padx=15, pady=15, fill=tk.X)
        
        section_title = tk.Label(button_inner, text="작업 선택", font=("Helvetica", 13, "bold"), bg='white', fg='#2d3748', anchor=tk.W)
        section_title.pack(fill=tk.X, pady=(0, 10))
        
        button_grid = tk.Frame(button_inner, bg='white')
        button_grid.pack(fill=tk.X)
        
        color_map = {
            "#673AB7": "#5E35B1", "#10b981": "#059669", "#3b82f6": "#2563eb",
            "#f59e0b": "#d97706", "#8b5cf6": "#7c3aed", "#009688": "#00796B"
        }
        
        def create_button(parent, text, color, command):
            btn = tk.Button(parent, text=text, command=command, font=("Helvetica", 11, "bold"),
                           bg=color, fg='white', relief=tk.FLAT, bd=0, padx=15, pady=12, cursor='hand2',
                           activebackground=color_map.get(color, color), activeforeground='white')
            dark = color_map.get(color, color)
            btn.bind("<Enter>", lambda e: btn.config(bg=dark))
            btn.bind("<Leave>", lambda e: btn.config(bg=color))
            self.buttons.append(btn)
            return btn
        
        create_button(button_grid, "AI 주제 추천", "#673AB7", self.recommend_topics).grid(row=0, column=0, padx=4, pady=4, sticky='nsew')
        create_button(button_grid, "엑셀 파일 생성", "#10b981", self.create_excel).grid(row=0, column=1, padx=4, pady=4, sticky='nsew')
        create_button(button_grid, "블로그 본문 생성", "#3b82f6", self.generate_content).grid(row=0, column=2, padx=4, pady=4, sticky='nsew')
        create_button(button_grid, "블로그 업로드", "#f59e0b", self.upload_blog).grid(row=1, column=0, padx=4, pady=4, sticky='nsew')
        create_button(button_grid, "전체 실행", "#8b5cf6", self.run_all).grid(row=1, column=1, padx=4, pady=4, sticky='nsew')
        create_button(button_grid, "워드+이미지", "#009688", self.create_word_doc).grid(row=1, column=2, padx=4, pady=4, sticky='nsew')
        
        for i in range(3):
            button_grid.columnconfigure(i, weight=1)
        
        schedule_card = tk.Frame(container, bg='white', relief=tk.FLAT, bd=0)
        schedule_card.pack(fill=tk.X, pady=(0, 12))
        
        schedule_inner = tk.Frame(schedule_card, bg='white')
        schedule_inner.pack(padx=15, pady=15, fill=tk.X)
        
        schedule_header = tk.Frame(schedule_inner, bg='white')
        schedule_header.pack(fill=tk.X, pady=(0, 8))
        
        tk.Label(schedule_header, text="예약 발행 설정", font=("Helvetica", 13, "bold"), bg='white', fg='#2d3748').pack(side=tk.LEFT)
        
        self.schedule_enabled = tk.BooleanVar(value=False)
        tk.Checkbutton(schedule_header, text="사용", variable=self.schedule_enabled, font=("Helvetica", 11), bg='white', command=self.toggle_schedule_options).pack(side=tk.LEFT, padx=(15, 0))
        
        self.schedule_options = tk.Frame(schedule_inner, bg='white')
        
        row1 = tk.Frame(self.schedule_options, bg='white')
        row1.pack(fill=tk.X, pady=4)
        tk.Label(row1, text="시작 날짜:", font=("Helvetica", 11), bg='white', width=10, anchor='w').pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.start_date = DateEntry(row1, width=12, font=("Helvetica", 11), background='#667eea', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        else:
            self.start_date = tk.Entry(row1, width=14, font=("Helvetica", 11))
            self.start_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.start_date.pack(side=tk.LEFT, padx=5)
        if not HAS_CALENDAR:
            tk.Label(row1, text="(YYYY-MM-DD)", font=("Helvetica", 10), bg='white', fg='#888').pack(side=tk.LEFT)
        
        row2 = tk.Frame(self.schedule_options, bg='white')
        row2.pack(fill=tk.X, pady=4)
        tk.Label(row2, text="시작 시간:", font=("Helvetica", 11), bg='white', width=10, anchor='w').pack(side=tk.LEFT)
        self.start_hour = ttk.Combobox(row2, values=[f"{i:02d}" for i in range(24)], width=5, font=("Helvetica", 11), state="readonly")
        self.start_hour.set("09")
        self.start_hour.pack(side=tk.LEFT, padx=2)
        tk.Label(row2, text="시", font=("Helvetica", 11), bg='white').pack(side=tk.LEFT, padx=(0, 8))
        self.start_minute = ttk.Combobox(row2, values=["00", "10", "20", "30", "40", "50"], width=5, font=("Helvetica", 11), state="readonly")
        self.start_minute.set("00")
        self.start_minute.pack(side=tk.LEFT, padx=2)
        tk.Label(row2, text="분", font=("Helvetica", 11), bg='white').pack(side=tk.LEFT)
        
        row3 = tk.Frame(self.schedule_options, bg='white')
        row3.pack(fill=tk.X, pady=4)
        tk.Label(row3, text="글 간격:", font=("Helvetica", 11), bg='white', width=10, anchor='w').pack(side=tk.LEFT)
        self.interval_minutes = tk.Spinbox(row3, from_=10, to=1440, increment=10, width=6, font=("Helvetica", 11))
        self.interval_minutes.delete(0, tk.END)
        self.interval_minutes.insert(0, "30")
        self.interval_minutes.pack(side=tk.LEFT, padx=2)
        tk.Label(row3, text="분 간격", font=("Helvetica", 11), bg='white').pack(side=tk.LEFT)
        
        progress_card = tk.Frame(container, bg='white', relief=tk.FLAT, bd=0)
        progress_card.pack(fill=tk.X, pady=(0, 12))
        
        progress_inner = tk.Frame(progress_card, bg='white')
        progress_inner.pack(padx=15, pady=15, fill=tk.X)
        
        tk.Label(progress_inner, text="진행 상황", font=("Helvetica", 13, "bold"), bg='white', fg='#2d3748', anchor=tk.W).pack(fill=tk.X, pady=(0, 6))
        
        self.progress_var = tk.StringVar(value="대기 중...")
        tk.Label(progress_inner, textvariable=self.progress_var, font=("Helvetica", 11), bg='white', fg='#4a5568', anchor=tk.W).pack(fill=tk.X, pady=(0, 6))
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Modern.Horizontal.TProgressbar", background='#667eea', troughcolor='#e2e8f0', borderwidth=0)
        self.progress_bar = ttk.Progressbar(progress_inner, mode='indeterminate', length=400, style='Modern.Horizontal.TProgressbar')
        self.progress_bar.pack(fill=tk.X)
        
        log_card = tk.Frame(container, bg='white', relief=tk.FLAT, bd=0)
        log_card.pack(fill=tk.BOTH, expand=True)
        
        log_inner = tk.Frame(log_card, bg='white')
        log_inner.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        log_header = tk.Frame(log_inner, bg='white')
        log_header.pack(fill=tk.X, pady=(0, 6))
        tk.Label(log_header, text="실행 로그", font=("Helvetica", 13, "bold"), bg='white', fg='#2d3748').pack(side=tk.LEFT)
        tk.Button(log_header, text="복사", command=self.copy_log, font=("Helvetica", 9), bg='#607D8B', fg='white', relief=tk.FLAT, padx=8).pack(side=tk.RIGHT, padx=2)
        tk.Button(log_header, text="지우기", command=self.clear_log, font=("Helvetica", 9), bg='#9E9E9E', fg='white', relief=tk.FLAT, padx=8).pack(side=tk.RIGHT, padx=2)
        
        self.log_text = scrolledtext.ScrolledText(log_inner, height=10, wrap=tk.WORD, font=("Courier", 10), bg='#1a202c', fg='#e2e8f0', relief=tk.FLAT, bd=0, padx=10, pady=10)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        status_frame = tk.Frame(root, bg='#2d3748', height=32)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        status_frame.pack_propagate(False)
        
        self.status_var = tk.StringVar(value="준비됨")
        tk.Label(status_frame, textvariable=self.status_var, anchor=tk.W, font=("Helvetica", 10), bg='#2d3748', fg='#cbd5e0', padx=15).pack(side=tk.LEFT, fill=tk.Y)
        tk.Label(status_frame, text="v1.1.0", font=("Helvetica", 9), bg='#2d3748', fg='#718096', padx=15).pack(side=tk.RIGHT, fill=tk.Y)
        
        self.log("애플리케이션이 시작되었습니다.")
        self.log("버튼을 클릭하여 작업을 시작하세요.")
    
    def toggle_schedule_options(self):
        if self.schedule_enabled.get():
            self.schedule_options.pack(fill=tk.X, pady=(8, 0))
        else:
            self.schedule_options.pack_forget()
    
    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        try:
            with open(self.log_file_path, "a", encoding="utf-8") as f:
                f.write(f"[{timestamp}] {message}\n")
        except:
            pass
    
    def copy_log(self):
        self.root.clipboard_clear()
        self.root.clipboard_append(self.log_text.get("1.0", tk.END))
        messagebox.showinfo("복사", "로그가 복사되었습니다.")
    
    def clear_log(self):
        self.log_text.delete("1.0", tk.END)
    
    def update_status(self, message):
        self.status_var.set(message)
        self.progress_var.set(message)
        self.root.update_idletasks()
    
    def set_buttons_state(self, state):
        for btn in self.buttons:
            btn.config(state=state)
    
    def run_script(self, script_name, description):
        def run():
            try:
                self.update_status(f"{description} 실행 중...")
                self.progress_bar.start(10)
                self.log(f"{description} 시작")
                
                base_path = get_base_path()
                script_path = os.path.join(base_path, script_name)
                if not os.path.exists(script_path):
                    script_path = script_name
                
                process = subprocess.Popen([get_python_executable(), script_path], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1, universal_newlines=True, cwd=base_path)
                
                for line in process.stdout:
                    if line.strip():
                        self.log(line.strip())
                
                process.wait()
                
                if process.returncode == 0:
                    self.update_status(f"{description} 완료")
                    self.log(f"{description} 완료!")
                    messagebox.showinfo("완료", f"{description}가 완료되었습니다.")
                else:
                    self.update_status(f"{description} 실패")
                    self.log(f"{description} 실패")
                    messagebox.showerror("오류", f"{description} 오류 발생")
            except Exception as e:
                self.update_status("오류")
                self.log(f"오류: {e}")
                messagebox.showerror("오류", str(e))
            finally:
                self.progress_bar.stop()
                self.set_buttons_state(tk.NORMAL)
        
        self.set_buttons_state(tk.DISABLED)
        threading.Thread(target=run, daemon=True).start()
    
    def recommend_topics(self):
        self.run_script("trend_analyzer.py", "AI 주제 추천")
    
    def create_excel(self):
        self.run_script("excel_create.py", "엑셀 파일 생성")
    
    def generate_content(self):
        self.run_script("create.py", "블로그 본문 생성")
    
    def create_word_doc(self):
        self.run_script("create_word.py", "워드+이미지 생성")
    
    def update_excel_schedule(self):
        if not self.schedule_enabled.get():
            return True
        try:
            base_path = get_base_path()
            files = glob.glob(os.path.join(base_path, "blog*.xlsx"))
            if not files:
                self.log("엑셀 파일 없음")
                return False
            
            excel_file = max(files, key=os.path.getmtime)
            self.log(f"예약 설정: {os.path.basename(excel_file)}")
            
            wb = load_workbook(excel_file)
            ws = wb.active
            
            if HAS_CALENDAR:
                start_date_obj = self.start_date.get_date()
            else:
                start_date_obj = datetime.strptime(self.start_date.get(), "%Y-%m-%d").date()
            
            start_datetime = datetime.combine(start_date_obj, datetime.strptime(f"{self.start_hour.get()}:{self.start_minute.get()}", "%H:%M").time())
            interval = int(self.interval_minutes.get())
            
            count = 0
            for row in range(2, ws.max_row + 1):
                if ws[f'A{row}'].value and ws[f'B{row}'].value:
                    t = start_datetime + timedelta(minutes=interval * count)
                    ws[f'D{row}'] = t.strftime("%Y-%m-%d %H:%M")
                    self.log(f"  {row}행: {t.strftime('%H:%M')}")
                    count += 1
            
            wb.save(excel_file)
            wb.close()
            self.log(f"{count}개 글 예약 완료")
            return True
        except Exception as e:
            self.log(f"예약 오류: {e}")
            return False
    
    def upload_blog(self):
        if self.schedule_enabled.get() and not self.update_excel_schedule():
            messagebox.showerror("오류", "예약 설정 실패")
            return
        self.run_script("upload_bot.py", "블로그 업로드")
    
    def run_all(self):
        def run_all_scripts():
            try:
                self.progress_bar.start(10)
                base_path = get_base_path()
                
                for script, desc in [("excel_create.py", "엑셀 생성"), ("create.py", "본문 생성")]:
                    self.update_status(f"{desc} 중...")
                    self.log(f"=== {desc} ===")
                    script_path = os.path.join(base_path, script)
                    if not os.path.exists(script_path):
                        script_path = script
                    process = subprocess.Popen([get_python_executable(), script_path], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1, universal_newlines=True, cwd=base_path)
                    for line in process.stdout:
                        if line.strip():
                            self.log(line.strip())
                    process.wait()
                    if process.returncode != 0:
                        raise Exception(f"{desc} 실패")
                
                if self.schedule_enabled.get():
                    self.log("=== 예약 설정 ===")
                    if not self.update_excel_schedule():
                        raise Exception("예약 실패")
                
                self.update_status("업로드 중...")
                self.log("=== 업로드 ===")
                script_path = os.path.join(base_path, "upload_bot.py")
                process = subprocess.Popen([get_python_executable(), script_path], stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1, universal_newlines=True, cwd=base_path)
                for line in process.stdout:
                    if line.strip():
                        self.log(line.strip())
                process.wait()
                if process.returncode != 0:
                    raise Exception("업로드 실패")
                
                self.update_status("완료!")
                self.log("모든 작업 완료!")
                messagebox.showinfo("완료", "모든 작업 완료!")
            except Exception as e:
                self.update_status("오류")
                self.log(f"오류: {e}")
                messagebox.showerror("오류", str(e))
            finally:
                self.progress_bar.stop()
                self.set_buttons_state(tk.NORMAL)
        
        self.set_buttons_state(tk.DISABLED)
        threading.Thread(target=run_all_scripts, daemon=True).start()

def main():
    root = tk.Tk()
    BlogAutomationGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
