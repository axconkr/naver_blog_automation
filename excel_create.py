from openpyxl import Workbook
from datetime import datetime
import os

# 현재 작업 디렉토리 확인
current_dir = os.getcwd()
print(f"현재 작업 디렉토리: {current_dir}")

# 날짜 형식: YYYYMMDD
date_str = datetime.now().strftime("%Y%m%d")
filename = f"blog{date_str}.xlsx"
filepath = os.path.join(current_dir, filename)

# 새 워크북 생성
wb = Workbook()
ws = wb.active

# 헤더 작성
ws['A1'] = "제목"
ws['B1'] = "본문"
ws['C1'] = "카테고리"
ws['D1'] = "발행시간"

# 블로그 포스팅 샘플 데이터 (제목, 카테고리, 발행시간)
blog_data = [
    ("파이썬 초보자를 위한 완벽 가이드 2026", "AI 활용법", ""),
    ("맛집 추천 베스트 10 - 서울 강남구 핫플레이스", "일상 기록", ""),
    ("집에서 할 수 있는 간단한 운동 루틴 5가지", "일상 기록", ""),
    ("주식 투자 시작하기 - 초보자를 위한 종목 추천", "일상 기록", ""),
    ("여행 준비 체크리스트 - 해외여행 필수 준비물", "일상 기록", ""),
]

for i, (title, category, schedule_time) in enumerate(blog_data, start=2):
    ws[f'A{i}'] = title
    ws[f'C{i}'] = category
    ws[f'D{i}'] = schedule_time

# 파일 저장
wb.save(filepath)
print(f"엑셀 파일 생성 완료: {filepath}")
print(f"파일명: {filename}")

# 워크북 닫기
wb.close()

print("작업 완료!")
