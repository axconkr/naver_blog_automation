"""
트렌드 분석 및 AI 주제 추천 모듈
"""

import requests
from bs4 import BeautifulSoup
import google.generativeai as genai
from openpyxl import Workbook
from dotenv import load_dotenv
from datetime import datetime
import os
import json
import re

load_dotenv()

NAVER_ID = os.getenv("NAVER_ID")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

if not GEMINI_API_KEY:
    raise ValueError(".env 파일에서 GEMINI_API_KEY를 찾을 수 없습니다.")

genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel('gemini-2.5-flash')


def get_naver_trending_keywords():
    """네이버 실시간 급상승 검색어 수집"""
    print("네이버 트렌드 키워드 수집 중...")
    
    keywords = []
    
    try:
        url = "https://www.naver.com/"
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
        }
        response = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        trend_items = soup.select('.ah_roll_area .ah_item, .ah_list .ah_item')
        for item in trend_items[:20]:
            keyword = item.select_one('.ah_k')
            if keyword:
                keywords.append(keyword.text.strip())
    except Exception as e:
        print(f"  네이버 메인 수집 실패: {e}")
    
    try:
        signal_url = "https://www.signal.bz/news"
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
        }
        response = requests.get(signal_url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        news_items = soup.select('.rank-text, .news-title, h3 a')
        for item in news_items[:20]:
            text = item.text.strip()
            if text and len(text) > 2:
                keywords.append(text)
    except Exception as e:
        print(f"  Signal.bz 수집 실패: {e}")
    
    try:
        zum_url = "https://zum.com/"
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
        }
        response = requests.get(zum_url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        issue_items = soup.select('.issue_keyword a, .keyword_item')
        for item in issue_items[:20]:
            text = item.text.strip()
            if text and len(text) > 2:
                keywords.append(text)
    except Exception as e:
        print(f"  ZUM 수집 실패: {e}")
    
    try:
        google_url = "https://trends.google.co.kr/trending?geo=KR"
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
        }
        response = requests.get(google_url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        trend_items = soup.select('.feed-item, .title a')
        for item in trend_items[:20]:
            text = item.text.strip()
            if text and len(text) > 2:
                keywords.append(text)
    except Exception as e:
        print(f"  Google Trends 수집 실패: {e}")
    
    unique_keywords = list(dict.fromkeys(keywords))
    print(f"  수집된 키워드: {len(unique_keywords)}개")
    
    return unique_keywords[:30]


def get_naver_news_headlines():
    """네이버 뉴스 헤드라인 수집"""
    print("네이버 뉴스 헤드라인 수집 중...")
    
    headlines = []
    categories = ['politics', 'economy', 'society', 'life', 'world', 'it']
    
    for category in categories[:4]:
        try:
            url = f"https://news.naver.com/section/{category}"
            headers = {
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
            }
            response = requests.get(url, headers=headers, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            news_items = soup.select('.sa_text_title, .cluster_text_headline')
            for item in news_items[:5]:
                text = item.text.strip()
                if text and len(text) > 10:
                    headlines.append({"category": category, "title": text})
        except Exception as e:
            print(f"  {category} 뉴스 수집 실패: {e}")
    
    print(f"  수집된 헤드라인: {len(headlines)}개")
    return headlines


def analyze_user_blog(blog_id):
    """사용자 블로그 분석"""
    print(f"블로그 분석 중: {blog_id}")
    
    blog_data = {
        "blog_id": blog_id,
        "titles": [],
        "categories": [],
        "keywords": []
    }
    
    try:
        url = f"https://blog.naver.com/PostList.naver?blogId={blog_id}&categoryNo=0&from=postList"
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
        }
        response = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        title_items = soup.select('.pcol2 .ell, .title a, .se-title-text')
        for item in title_items[:20]:
            text = item.text.strip()
            if text and len(text) > 3:
                blog_data["titles"].append(text)
        
        cat_items = soup.select('.category a, .link_category')
        for item in cat_items[:10]:
            text = item.text.strip()
            if text:
                blog_data["categories"].append(text)
                
    except Exception as e:
        print(f"  블로그 분석 실패: {e}")
    
    try:
        rss_url = f"https://rss.blog.naver.com/{blog_id}.xml"
        response = requests.get(rss_url, timeout=10)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'xml')
            items = soup.find_all('item')
            for item in items[:20]:
                title = item.find('title')
                if title:
                    blog_data["titles"].append(title.text.strip())
    except Exception as e:
        print(f"  RSS 분석 실패: {e}")
    
    blog_data["titles"] = list(dict.fromkeys(blog_data["titles"]))[:20]
    blog_data["categories"] = list(dict.fromkeys(blog_data["categories"]))[:10]
    
    print(f"  분석된 글 제목: {len(blog_data['titles'])}개")
    print(f"  발견된 카테고리: {len(blog_data['categories'])}개")
    
    return blog_data


def generate_topic_recommendations(trends, news, blog_data, num_topics=10):
    """AI 기반 주제 추천"""
    print("AI 주제 추천 생성 중...")
    
    prompt = f"""당신은 네이버 블로그 SEO 전문가입니다.

## 현재 트렌드 키워드
{', '.join(trends[:20]) if trends else '수집된 키워드 없음'}

## 최신 뉴스 헤드라인
{json.dumps(news[:15], ensure_ascii=False, indent=2) if news else '수집된 뉴스 없음'}

## 사용자 블로그 분석 결과
- 블로그 ID: {blog_data.get('blog_id', 'unknown')}
- 기존 글 제목들: {', '.join(blog_data.get('titles', [])[:10]) if blog_data.get('titles') else '분석된 글 없음'}
- 카테고리: {', '.join(blog_data.get('categories', [])) if blog_data.get('categories') else '카테고리 정보 없음'}

## 요청
위 정보를 바탕으로 이 블로그에 가장 적합한 블로그 글 주제 {num_topics}개를 추천해주세요.

## 추천 기준
1. 현재 트렌드와 관련성 높은 주제
2. 사용자 블로그 스타일/분야에 맞는 주제
3. 네이버 검색 상위 노출 가능성 높은 주제
4. 클릭을 유도하는 매력적인 제목

## 출력 형식 (JSON)
정확히 아래 형식으로만 출력하세요. 다른 텍스트 없이 JSON만 출력:
{{
  "recommendations": [
    {{
      "title": "블로그 제목 (35자 이내)",
      "category": "카테고리",
      "trend_keyword": "관련 트렌드 키워드",
      "reason": "추천 이유 (1문장)"
    }}
  ]
}}
"""

    try:
        response = model.generate_content(prompt)
        text = response.text.strip()
        
        json_match = re.search(r'\{[\s\S]*\}', text)
        if json_match:
            result = json.loads(json_match.group())
            recommendations = result.get("recommendations", [])
            print(f"  추천 주제: {len(recommendations)}개 생성")
            return recommendations
        else:
            print(f"  JSON 파싱 실패, 텍스트 응답 처리")
            return []
            
    except Exception as e:
        print(f"  AI 추천 생성 실패: {e}")
        return []


def save_recommendations_to_excel(recommendations, output_path=None):
    """추천 주제를 엑셀 파일로 저장"""
    if output_path is None:
        today = datetime.now().strftime("%Y%m%d")
        output_path = os.path.join(os.path.dirname(__file__), f"blog{today}.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "블로그 주제"
    
    ws['A1'] = '제목'
    ws['B1'] = '본문'
    ws['C1'] = '카테고리'
    ws['D1'] = '예약시간'
    ws['E1'] = '트렌드키워드'
    ws['F1'] = '추천이유'
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 40
    
    for i, rec in enumerate(recommendations, start=2):
        ws[f'A{i}'] = rec.get('title', '')
        ws[f'C{i}'] = rec.get('category', '')
        ws[f'E{i}'] = rec.get('trend_keyword', '')
        ws[f'F{i}'] = rec.get('reason', '')
    
    wb.save(output_path)
    print(f"\n✓ 엑셀 저장 완료: {output_path}")
    
    return output_path


def recommend_topics(num_topics=10):
    """메인 함수: 트렌드 분석 → 블로그 분석 → AI 추천 → 엑셀 저장"""
    print("=" * 50)
    print("트렌드 기반 AI 주제 추천 시스템")
    print("=" * 50 + "\n")
    
    trends = get_naver_trending_keywords()
    
    news = get_naver_news_headlines()
    
    blog_data = {}
    if NAVER_ID:
        blog_data = analyze_user_blog(NAVER_ID)
    else:
        print("⚠️ .env에 NAVER_ID가 없어 블로그 분석을 건너뜁니다.")
    
    print()
    recommendations = generate_topic_recommendations(trends, news, blog_data, num_topics)
    
    if recommendations:
        output_path = save_recommendations_to_excel(recommendations)
        
        print("\n" + "=" * 50)
        print("추천된 블로그 주제")
        print("=" * 50)
        for i, rec in enumerate(recommendations, 1):
            print(f"\n{i}. {rec.get('title', 'N/A')}")
            print(f"   카테고리: {rec.get('category', 'N/A')}")
            print(f"   트렌드: {rec.get('trend_keyword', 'N/A')}")
            print(f"   이유: {rec.get('reason', 'N/A')}")
        
        return output_path
    else:
        print("\n✗ 주제 추천 생성 실패")
        return None


if __name__ == "__main__":
    result = recommend_topics(num_topics=10)
    if result:
        print(f"\n완료! 엑셀 파일을 확인하세요: {result}")
